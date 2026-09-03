from collections import defaultdict
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, send_file

from flask_sqlalchemy import SQLAlchemy

from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    login_required,
    current_user
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///finance.db"
app.config["SECRET_KEY"] = "mysecretkey"

db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


# ============================================================
# MODELS
# ============================================================

class User(UserMixin, db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    username = db.Column(
        db.String(100),
        nullable=False
    )

    email = db.Column(
        db.String(120),
        unique=True,
        nullable=False
    )

    password = db.Column(
        db.String(200),
        nullable=False
    )


class Expense(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    title = db.Column(
        db.String(100),
        nullable=False
    )

    category = db.Column(
        db.String(50),
        nullable=False
    )

    amount = db.Column(
        db.Float,
        nullable=False
    )

    date = db.Column(
        db.String(20),
        nullable=False
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey("user.id")
    )


class Budget(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    amount = db.Column(
        db.Float,
        nullable=False
    )

    month = db.Column(
        db.String(20),
        nullable=False
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey("user.id")
    )


# ============================================================
# FINANCIAL GOAL MODEL
# ============================================================

class FinancialGoal(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    goal_name = db.Column(
        db.String(100),
        nullable=False
    )

    target_amount = db.Column(
        db.Float,
        nullable=False
    )

    current_amount = db.Column(
        db.Float,
        nullable=False,
        default=0
    )

    target_date = db.Column(
        db.String(20),
        nullable=False
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey("user.id"),
        nullable=False
    )


# ============================================================
# INVESTMENT MODEL
# ============================================================

class Investment(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    asset_name = db.Column(
        db.String(100),
        nullable=False
    )

    asset_type = db.Column(
        db.String(50),
        nullable=False
    )

    quantity = db.Column(
        db.Float,
        nullable=False
    )

    purchase_price = db.Column(
        db.Float,
        nullable=False
    )

    current_price = db.Column(
        db.Float,
        nullable=False
    )

    purchase_date = db.Column(
        db.String(20),
        nullable=False
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey("user.id"),
        nullable=False
    )

    @property
    def invested_amount(self):

        return (
            self.quantity *
            self.purchase_price
        )

    @property
    def current_value(self):

        return (
            self.quantity *
            self.current_price
        )

    @property
    def profit_loss(self):

        return (
            self.current_value -
            self.invested_amount
        )

    @property
    def return_percentage(self):

        if self.invested_amount == 0:
            return 0

        return (
            self.profit_loss /
            self.invested_amount
        ) * 100


# ============================================================
# ADD FINANCIAL GOAL
# ============================================================

@app.route(
    "/add_goal",
    methods=["GET", "POST"]
)
@login_required
def add_goal():

    if request.method == "POST":

        goal_name = request.form["goal_name"]

        target_amount = float(
            request.form["target_amount"]
        )

        current_amount = float(
            request.form["current_amount"]
        )

        target_date = request.form["target_date"]

        goal = FinancialGoal(

            goal_name=goal_name,

            target_amount=target_amount,

            current_amount=current_amount,

            target_date=target_date,

            user_id=current_user.id
        )

        db.session.add(goal)

        db.session.commit()

        return redirect(
            url_for("goals")
        )

    return render_template(
        "add_goal.html"
    )


# ============================================================
# VIEW FINANCIAL GOALS
# ============================================================

@app.route("/goals")
@login_required
def goals():

    goals = FinancialGoal.query.filter_by(
        user_id=current_user.id
    ).all()

    return render_template(
        "goals.html",
        goals=goals
    )


# ============================================================
# EDIT FINANCIAL GOAL
# ============================================================

@app.route(
    "/edit_goal/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_goal(id):

    goal = FinancialGoal.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first_or_404()

    if request.method == "POST":

        goal.goal_name = request.form["goal_name"]

        goal.target_amount = float(
            request.form["target_amount"]
        )

        goal.current_amount = float(
            request.form["current_amount"]
        )

        goal.target_date = request.form["target_date"]

        db.session.commit()

        return redirect(
            url_for("goals")
        )

    return render_template(
        "edit_goal.html",
        goal=goal
    )


# ============================================================
# DELETE FINANCIAL GOAL
# ============================================================

@app.route(
    "/delete_goal/<int:id>"
)
@login_required
def delete_goal(id):

    goal = FinancialGoal.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first_or_404()

    db.session.delete(goal)

    db.session.commit()

    return redirect(
        url_for("goals")
    )


# ============================================================
# LOGIN MANAGER
# ============================================================

@login_manager.user_loader
def load_user(user_id):

    return User.query.get(
        int(user_id)
    )


# ============================================================
# LOGIN
# ============================================================

@app.route(
    "/",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        email = request.form["email"]

        password = request.form["password"]

        user = User.query.filter_by(
            email=email,
            password=password
        ).first()

        if user:

            login_user(user)

            return redirect(
                url_for("dashboard")
            )

        else:

            return "Invalid Email or Password"

    return render_template(
        "login.html"
    )


# ============================================================
# REGISTER
# ============================================================

@app.route(
    "/register",
    methods=["GET", "POST"]
)
def register():

    if request.method == "POST":

        username = request.form["username"]

        email = request.form["email"]

        password = request.form["password"]

        new_user = User(
            username=username,
            email=email,
            password=password
        )

        db.session.add(new_user)

        db.session.commit()

        return redirect(
            url_for("login")
        )

    return render_template(
        "register.html"
    )


# ============================================================
# FINANCIAL HEALTH SCORE ENGINE
# ============================================================

def calculate_financial_health_score(user_id):

    expenses = Expense.query.filter_by(
        user_id=user_id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=user_id
    ).all()

    goals = FinancialGoal.query.filter_by(
        user_id=user_id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    # Savings Rate Score

    if total_budget > 0:

        savings_rate = (
            total_budget -
            total_expenses
        ) / total_budget

        savings_score = max(
            0,
            min(
                40,
                savings_rate * 40
            )
        )

    else:

        savings_score = 0

    # Budget Adherence Score

    if total_budget > 0:

        usage_percentage = (
            total_expenses /
            total_budget
        ) * 100

        if usage_percentage <= 100:

            adherence_score = 30

        elif usage_percentage <= 120:

            adherence_score = 15

        else:

            adherence_score = 0

    else:

        usage_percentage = 0
        adherence_score = 0

    # Goal Progress Score

    if goals:

        progress_values = []

        for goal in goals:

            if goal.target_amount > 0:

                progress = (
                    goal.current_amount /
                    goal.target_amount
                ) * 100

                progress_values.append(
                    min(progress, 100)
                )

        average_progress = (

            sum(progress_values) /
            len(progress_values)

            if progress_values

            else 0
        )

        goal_score = (
            average_progress /
            100
        ) * 30

    else:

        average_progress = 0
        goal_score = 0

    total_score = round(
        savings_score +
        adherence_score +
        goal_score,
        2
    )

    if total_score >= 80:

        rating = "Excellent"

    elif total_score >= 60:

        rating = "Good"

    elif total_score >= 40:

        rating = "Fair"

    else:

        rating = "Needs Improvement"

    return {

        "score": total_score,

        "rating": rating,

        "savings_score": round(
            savings_score,
            2
        ),

        "adherence_score": round(
            adherence_score,
            2
        ),

        "goal_score": round(
            goal_score,
            2
        ),

        "usage_percentage": round(
            usage_percentage,
            2
        ),

        "average_goal_progress": round(
            average_progress,
            2
        )
    }


# ============================================================
# FINANCIAL HEALTH PAGE
# ============================================================

@app.route("/financial_health")
@login_required
def financial_health():

    health_score = calculate_financial_health_score(
        current_user.id
    )

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    savings = total_budget - total_expenses

    return render_template(
        "financial_health.html",
        health_score=health_score["score"],
        health_status=health_score["rating"],
        total_budget=total_budget,
        total_expenses=total_expenses,
        savings=savings,
        expense_ratio=health_score["usage_percentage"],
        expense_score=health_score["adherence_score"],
        savings_score=health_score["savings_score"],
        goal_score=health_score["goal_score"],
        goal_progress=health_score["average_goal_progress"]
    )


# ============================================================
# INTEGRATION HELPERS
# ============================================================

def get_quick_alerts(
    user_id,
    limit=3
):

    expenses = Expense.query.filter_by(
        user_id=user_id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=user_id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    quick_alerts = []

    if total_budget > 0:

        usage_percentage = (
            total_expenses /
            total_budget
        ) * 100

        if usage_percentage > 100:

            quick_alerts.append({

                "type": "danger",

                "message":
                    f"Budget exceeded by "
                    f"₹{total_expenses - total_budget:.2f}."
            })

        elif usage_percentage >= 90:

            quick_alerts.append({

                "type": "warning",

                "message":
                    f"{usage_percentage:.2f}% "
                    f"of your budget is used."
            })

        elif usage_percentage >= 75:

            quick_alerts.append({

                "type": "info",

                "message":
                    f"{usage_percentage:.2f}% "
                    f"of your budget is used."
            })

    else:

        quick_alerts.append({

            "type": "info",

            "message":
                "No budget set for this month."
        })

    category_spending = {}

    for expense in expenses:

        category_spending[
            expense.category
        ] = (

            category_spending.get(
                expense.category,
                0
            )

            + expense.amount
        )

    if category_spending:

        top_category = max(
            category_spending,
            key=category_spending.get
        )

        quick_alerts.append({

            "type": "warning",

            "message":
                f"Highest spending category: "
                f"{top_category} "
                f"(₹{category_spending[top_category]:.2f})."
        })

    return quick_alerts[:limit]


# ============================================================
# TOP BUDGET RECOMMENDATION
# ============================================================

def get_top_budget_recommendation(user_id):

    expenses = Expense.query.filter_by(
        user_id=user_id
    ).all()

    category_data = {}

    for expense in expenses:

        category_data.setdefault(
            expense.category,
            []
        ).append(
            expense.amount
        )

    if not category_data:

        return None

    top_category = max(

        category_data,

        key=lambda c:
            sum(category_data[c])
    )

    amounts = category_data[
        top_category
    ]

    average = (
        sum(amounts) /
        len(amounts)
    )

    recommended = (
        average *
        0.90
    )

    return {

        "category": top_category,

        "average": round(
            average,
            2
        ),

        "recommended": round(
            recommended,
            2
        )
    }


# ============================================================
# SPENDING INSIGHT
# ============================================================

def get_spending_insight(user_id):

    expenses = Expense.query.filter_by(
        user_id=user_id
    ).all()

    if not expenses:

        return None

    monthly_totals = {}

    for expense in expenses:

        month_key = str(
            expense.date
        )[:7]

        monthly_totals[
            month_key
        ] = (

            monthly_totals.get(
                month_key,
                0
            )

            + expense.amount
        )

    sorted_months = sorted(
        monthly_totals.keys()
    )

    trend_direction = None
    trend_percentage = 0

    if len(sorted_months) >= 2:

        current_month = monthly_totals[
            sorted_months[-1]
        ]

        previous_month = monthly_totals[
            sorted_months[-2]
        ]

        if previous_month > 0:

            trend_percentage = (

                (
                    current_month -
                    previous_month
                )

                / previous_month

            ) * 100

            trend_direction = (

                "up"

                if trend_percentage > 0

                else "down"
            )

    return {

        "trend_direction":
            trend_direction,

        "trend_percentage":
            round(
                abs(trend_percentage),
                2
            )
    }


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/dashboard")
@login_required
def dashboard():

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    remaining_balance = (
        total_budget -
        total_expenses
    )

    # Pie Chart

    category_data = {}

    for expense in expenses:

        if expense.category in category_data:

            category_data[
                expense.category
            ] += expense.amount

        else:

            category_data[
                expense.category
            ] = expense.amount

    labels = list(
        category_data.keys()
    )

    values = list(
        category_data.values()
    )

    # Monthly Chart

    monthly_data = {}

    for expense in expenses:

        month = str(
            expense.date
        )[:7]

        if month in monthly_data:

            monthly_data[
                month
            ] += expense.amount

        else:

            monthly_data[
                month
            ] = expense.amount

    month_labels = list(
        monthly_data.keys()
    )

    month_values = list(
        monthly_data.values()
    )

    # Recent Transactions

    recent_expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Expense.id.desc()
    ).limit(5).all()

    # Integration Layer

    health_score = calculate_financial_health_score(
        current_user.id
    )

    quick_alerts = get_quick_alerts(
        current_user.id
    )

    top_recommendation = get_top_budget_recommendation(
        current_user.id
    )

    spending_insight = get_spending_insight(
        current_user.id
    )

    return render_template(

        "dashboard.html",

        total_budget=total_budget,

        total_expenses=total_expenses,

        remaining_balance=remaining_balance,

        labels=labels,

        values=values,

        month_labels=month_labels,

        month_values=month_values,

        recent_expenses=recent_expenses,

        health_score=health_score,

        quick_alerts=quick_alerts,

        top_recommendation=top_recommendation,

        spending_insight=spending_insight
    )


# ============================================================
# PROFILE
# ============================================================

@app.route(
    "/profile",
    methods=["GET", "POST"]
)
@login_required
def profile():

    if request.method == "POST":

        current_user.username = (
            request.form["username"]
        )

        current_user.email = (
            request.form["email"]
        )

        db.session.commit()

        return redirect(
            url_for("profile")
        )

    return render_template(
        "profile.html"
    )


# ============================================================
# ADD EXPENSE
# ============================================================

@app.route(
    "/add_expense",
    methods=["GET", "POST"]
)
@login_required
def add_expense():

    if request.method == "POST":

        expense = Expense(

            title=request.form["title"],

            category=request.form["category"],

            amount=float(
                request.form["amount"]
            ),

            date=request.form["date"],

            user_id=current_user.id
        )

        db.session.add(expense)

        db.session.commit()

        return redirect(
            url_for("expenses")
        )

    return render_template(
        "add_expense.html"
    )


# ============================================================
# VIEW EXPENSES
# ============================================================

@app.route("/expenses")
@login_required
def expenses():

    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()

    expense_query = Expense.query.filter_by(
        user_id=current_user.id
    )

    if search:
        expense_query = expense_query.filter(
            Expense.title.ilike(f"%{search}%")
        )

    if category:
        expense_query = expense_query.filter_by(
            category=category
        )

    expenses = expense_query.order_by(
        Expense.id.desc()
    ).all()

    return render_template(
        "expenses.html",
        expenses=expenses,
        search=search,
        category=category
    )


# ============================================================
# EDIT EXPENSE
# ============================================================

@app.route(
    "/edit_expense/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_expense(id):

    expense = Expense.query.get_or_404(
        id
    )

    if request.method == "POST":

        expense.title = (
            request.form["title"]
        )

        expense.category = (
            request.form["category"]
        )

        expense.amount = float(
            request.form["amount"]
        )

        expense.date = (
            request.form["date"]
        )

        db.session.commit()

        return redirect(
            url_for("expenses")
        )

    return render_template(
        "edit_expense.html",
        expense=expense
    )


# ============================================================
# DELETE EXPENSE
# ============================================================

@app.route(
    "/delete_expense/<int:id>"
)
@login_required
def delete_expense(id):

    expense = Expense.query.get_or_404(
        id
    )

    db.session.delete(
        expense
    )

    db.session.commit()

    return redirect(
        url_for("expenses")
    )


# ============================================================
# SPENDING PATTERN ANALYSIS
# ============================================================

@app.route("/spending_analysis")
@login_required
def spending_analysis():

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    if not expenses:

        return render_template(
            "spending_analysis.html",
            has_data=False
        )

    category_totals = defaultdict(float)

    for expense in expenses:

        category_totals[
            expense.category
        ] += expense.amount

    total_spent = sum(
        category_totals.values()
    )

    category_percentage = {

        cat:
            (
                amt /
                total_spent
            ) * 100

        for cat, amt
        in category_totals.items()
    }

    top_category = max(
        category_totals,
        key=category_totals.get
    )

    monthly_totals = defaultdict(float)

    for expense in expenses:

        month_key = str(
            expense.date
        )[:7]

        monthly_totals[
            month_key
        ] += expense.amount

    sorted_months = sorted(
        monthly_totals.keys()
    )

    trend_direction = None
    trend_percentage = 0

    if len(sorted_months) >= 2:

        current_month = monthly_totals[
            sorted_months[-1]
        ]

        previous_month = monthly_totals[
            sorted_months[-2]
        ]

        if previous_month > 0:

            trend_percentage = (

                (
                    current_month -
                    previous_month
                )

                / previous_month

            ) * 100

            trend_direction = (

                "up"

                if trend_percentage > 0

                else "down"
            )

    weekday_totals = defaultdict(float)

    weekday_names = [

        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday"
    ]

    for expense in expenses:

        try:

            parsed_date = datetime.strptime(
                expense.date,
                "%Y-%m-%d"
            )

            weekday_totals[
                weekday_names[
                    parsed_date.weekday()
                ]
            ] += expense.amount

        except ValueError:

            continue

    top_weekday = (

        max(
            weekday_totals,
            key=weekday_totals.get
        )

        if weekday_totals

        else None
    )

    highest_expense = max(
        expenses,
        key=lambda e: e.amount
    )

    average_expense = (
        total_spent /
        len(expenses)
    )

    return render_template(

        "spending_analysis.html",

        has_data=True,

        total_spent=total_spent,

        category_totals=dict(
            category_totals
        ),

        category_percentage=(
            category_percentage
        ),

        top_category=top_category,

        monthly_totals=dict(
            monthly_totals
        ),

        sorted_months=sorted_months,

        trend_direction=(
            trend_direction
        ),

        trend_percentage=round(
            abs(trend_percentage),
            2
        ),

        weekday_totals=dict(
            weekday_totals
        ),

        top_weekday=top_weekday,

        highest_expense=highest_expense,

        average_expense=round(
            average_expense,
            2
        )
    )


# ============================================================
# BUDGET
# ============================================================

@app.route(
    "/set_budget",
    methods=["GET", "POST"]
)
@login_required
def set_budget():

    if request.method == "POST":

        month = request.form["month"]

        amount = request.form["amount"]

        budget = Budget.query.filter_by(

            user_id=current_user.id,

            month=month

        ).first()

        if budget:

            budget.amount = amount

        else:

            budget = Budget(

                month=month,

                amount=amount,

                user_id=current_user.id
            )

            db.session.add(
                budget
            )

        db.session.commit()

        return redirect(
            url_for("dashboard")
        )

    return render_template(
        "set_budget.html"
    )


# ============================================================
# BUDGET RECOMMENDATION
# ============================================================

@app.route(
    "/budget-recommendation"
)
@login_required
def budget_recommendation():

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    category_data = {}

    for expense in expenses:

        category = expense.category

        if category not in category_data:

            category_data[
                category
            ] = []

        category_data[
            category
        ].append(
            expense.amount
        )

    recommendations = {}

    for category, amounts in category_data.items():

        total = sum(amounts)

        count = len(amounts)

        average = total / count

        recommended_budget = (
            average * 0.90
        )

        recommendations[
            category
        ] = {

            "average": round(
                average,
                2
            ),

            "recommended": round(
                recommended_budget,
                2
            )
        }

    return render_template(

        "budget_recommendation.html",

        recommendations=recommendations
    )


# ============================================================
# ALERTS & NOTIFICATIONS
# ============================================================

@app.route("/alerts")
@login_required
def alerts():

    alerts = []

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    if total_budget > 0:

        usage_percentage = (
            total_expenses /
            total_budget
        ) * 100

        if usage_percentage > 100:

            alerts.append({

                "type": "danger",

                "title":
                    "Budget Exceeded",

                "message":
                    f"You have exceeded "
                    f"your budget by "
                    f"₹{total_expenses - total_budget:.2f}."
            })

        elif usage_percentage >= 90:

            alerts.append({

                "type": "warning",

                "title":
                    "Budget Almost Exceeded",

                "message":
                    f"You have used "
                    f"{usage_percentage:.2f}% "
                    f"of your total budget."
            })

        elif usage_percentage >= 75:

            alerts.append({

                "type": "info",

                "title":
                    "Budget Alert",

                "message":
                    f"You have used "
                    f"{usage_percentage:.2f}% "
                    f"of your total budget."
            })

        else:

            alerts.append({

                "type": "success",

                "title":
                    "Budget Status",

                "message":
                    f"Your spending is within "
                    f"budget. You have used "
                    f"{usage_percentage:.2f}% "
                    f"of your budget."
            })

    else:

        alerts.append({

            "type": "info",

            "title":
                "No Budget Set",

            "message":
                "Please set a budget to receive "
                "budget alerts."
        })

    category_spending = {}

    for expense in expenses:

        category = expense.category

        if category not in category_spending:

            category_spending[
                category
            ] = 0

        category_spending[
            category
        ] += expense.amount

    if category_spending:

        highest_category = max(

            category_spending,

            key=category_spending.get
        )

        highest_amount = category_spending[
            highest_category
        ]

        alerts.append({

            "type": "warning",

            "title":
                "Highest Spending Category",

            "message":
                f"Your highest spending category "
                f"is {highest_category} with "
                f"spending of ₹{highest_amount:.2f}."
        })

    return render_template(
        "alerts.html",
        alerts=alerts
    )


# ============================================================
# ADD INVESTMENT
# ============================================================

@app.route(
    "/add_investment",
    methods=["GET", "POST"]
)
@login_required
def add_investment():

    if request.method == "POST":

        asset_name = request.form[
            "asset_name"
        ]

        asset_type = request.form[
            "asset_type"
        ]

        quantity = float(
            request.form["quantity"]
        )

        purchase_price = float(
            request.form["purchase_price"]
        )

        current_price = float(
            request.form["current_price"]
        )

        purchase_date = request.form[
            "purchase_date"
        ]

        investment = Investment(

            asset_name=asset_name,

            asset_type=asset_type,

            quantity=quantity,

            purchase_price=purchase_price,

            current_price=current_price,

            purchase_date=purchase_date,

            user_id=current_user.id
        )

        db.session.add(
            investment
        )

        db.session.commit()

        return redirect(
            url_for("investments")
        )

    return render_template(
        "add_investment.html"
    )


# ============================================================
# INVESTMENT PORTFOLIO
# ============================================================

@app.route("/investments")
@login_required
def investments():

    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()

    total_invested = sum(

        investment.invested_amount

        for investment in investments
    )

    total_value = sum(

        investment.current_value

        for investment in investments
    )

    total_profit_loss = (
        total_value -
        total_invested
    )

    if total_invested > 0:

        total_return = (

            total_profit_loss /
            total_invested

        ) * 100

    else:

        total_return = 0

    allocation = {}

    for investment in investments:

        asset_type = (
            investment.asset_type
        )

        if asset_type in allocation:

            allocation[
                asset_type
            ] += investment.current_value

        else:

            allocation[
                asset_type
            ] = investment.current_value

    allocation_percentage = {}

    for asset_type, value in allocation.items():

        if total_value > 0:

            allocation_percentage[
                asset_type
            ] = (

                value /
                total_value

            ) * 100

        else:

            allocation_percentage[
                asset_type
            ] = 0

    return render_template(

        "investments.html",

        investments=investments,

        total_invested=total_invested,

        total_value=total_value,

        total_profit_loss=(
            total_profit_loss
        ),

        total_return=total_return,

        allocation=allocation,

        allocation_percentage=(
            allocation_percentage
        )
    )


# ============================================================
# DELETE INVESTMENT
# ============================================================

@app.route(
    "/delete_investment/<int:id>"
)
@login_required
def delete_investment(id):

    investment = Investment.query.filter_by(

        id=id,

        user_id=current_user.id

    ).first_or_404()

    db.session.delete(
        investment
    )

    db.session.commit()

    return redirect(
        url_for("investments")
    )


# ============================================================
# EXCEL FINANCIAL REPORT
# ============================================================

@app.route("/download_excel")
@login_required
def download_excel():

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()

    goals = FinancialGoal.query.filter_by(
        user_id=current_user.id
    ).all()

    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    remaining_balance = (
        total_budget -
        total_expenses
    )

    total_invested = sum(
        investment.invested_amount
        for investment in investments
    )

    total_investment_value = sum(
        investment.current_value
        for investment in investments
    )

    total_profit_loss = (
        total_investment_value -
        total_invested
    )

    workbook = Workbook()

    # --------------------------------------------------------
    # Summary Sheet
    # --------------------------------------------------------

    summary = workbook.active

    summary.title = "Summary"

    summary["A1"] = (
        "FinSight Financial Report"
    )

    summary["A1"].font = Font(
        bold=True,
        size=16
    )

    summary["A3"] = (
        "Financial Summary"
    )

    summary["A3"].font = Font(
        bold=True
    )

    summary["A4"] = "Total Budget"
    summary["B4"] = total_budget

    summary["A5"] = "Total Expenses"
    summary["B5"] = total_expenses

    summary["A6"] = "Remaining Balance"
    summary["B6"] = remaining_balance

    summary["A7"] = "Total Invested"
    summary["B7"] = total_invested

    summary["A8"] = (
        "Current Investment Value"
    )

    summary["B8"] = (
        total_investment_value
    )

    summary["A9"] = (
        "Investment Profit/Loss"
    )

    summary["B9"] = (
        total_profit_loss
    )

    # --------------------------------------------------------
    # Expenses Sheet
    # --------------------------------------------------------

    expense_sheet = workbook.create_sheet(
        "Expenses"
    )

    expense_headers = [

        "ID",
        "Title",
        "Category",
        "Amount",
        "Date"
    ]

    for col, header in enumerate(
        expense_headers,
        1
    ):

        cell = expense_sheet.cell(

            row=1,

            column=col,

            value=header
        )

        cell.font = Font(
            bold=True
        )

    for row, expense in enumerate(
        expenses,
        2
    ):

        expense_sheet.cell(
            row,
            1,
            expense.id
        )

        expense_sheet.cell(
            row,
            2,
            expense.title
        )

        expense_sheet.cell(
            row,
            3,
            expense.category
        )

        expense_sheet.cell(
            row,
            4,
            expense.amount
        )

        expense_sheet.cell(
            row,
            5,
            expense.date
        )

    # --------------------------------------------------------
    # Budgets Sheet
    # --------------------------------------------------------

    budget_sheet = workbook.create_sheet(
        "Budgets"
    )

    budget_headers = [

        "ID",
        "Month",
        "Amount"
    ]

    for col, header in enumerate(
        budget_headers,
        1
    ):

        cell = budget_sheet.cell(

            row=1,

            column=col,

            value=header
        )

        cell.font = Font(
            bold=True
        )

    for row, budget in enumerate(
        budgets,
        2
    ):

        budget_sheet.cell(
            row,
            1,
            budget.id
        )

        budget_sheet.cell(
            row,
            2,
            budget.month
        )

        budget_sheet.cell(
            row,
            3,
            budget.amount
        )

    # --------------------------------------------------------
    # Goals Sheet
    # --------------------------------------------------------

    goal_sheet = workbook.create_sheet(
        "Goals"
    )

    goal_headers = [

        "ID",
        "Goal Name",
        "Target Amount",
        "Current Amount",
        "Target Date",
        "Progress %"
    ]

    for col, header in enumerate(
        goal_headers,
        1
    ):

        cell = goal_sheet.cell(

            row=1,

            column=col,

            value=header
        )

        cell.font = Font(
            bold=True
        )

    for row, goal in enumerate(
        goals,
        2
    ):

        if goal.target_amount > 0:

            progress = (

                goal.current_amount /
                goal.target_amount

            ) * 100

        else:

            progress = 0

        goal_sheet.cell(
            row,
            1,
            goal.id
        )

        goal_sheet.cell(
            row,
            2,
            goal.goal_name
        )

        goal_sheet.cell(
            row,
            3,
            goal.target_amount
        )

        goal_sheet.cell(
            row,
            4,
            goal.current_amount
        )

        goal_sheet.cell(
            row,
            5,
            goal.target_date
        )

        goal_sheet.cell(
            row,
            6,
            round(
                progress,
                2
            )
        )

    # --------------------------------------------------------
    # Investments Sheet
    # --------------------------------------------------------

    investment_sheet = workbook.create_sheet(
        "Investments"
    )

    investment_headers = [

        "ID",
        "Asset Name",
        "Asset Type",
        "Quantity",
        "Purchase Price",
        "Current Price",
        "Purchase Date",
        "Invested Amount",
        "Current Value",
        "Profit/Loss",
        "Return %"
    ]

    for col, header in enumerate(
        investment_headers,
        1
    ):

        cell = investment_sheet.cell(

            row=1,

            column=col,

            value=header
        )

        cell.font = Font(
            bold=True
        )

    for row, investment in enumerate(
        investments,
        2
    ):

        investment_sheet.cell(
            row,
            1,
            investment.id
        )

        investment_sheet.cell(
            row,
            2,
            investment.asset_name
        )

        investment_sheet.cell(
            row,
            3,
            investment.asset_type
        )

        investment_sheet.cell(
            row,
            4,
            investment.quantity
        )

        investment_sheet.cell(
            row,
            5,
            investment.purchase_price
        )

        investment_sheet.cell(
            row,
            6,
            investment.current_price
        )

        investment_sheet.cell(
            row,
            7,
            investment.purchase_date
        )

        investment_sheet.cell(
            row,
            8,
            investment.invested_amount
        )

        investment_sheet.cell(
            row,
            9,
            investment.current_value
        )

        investment_sheet.cell(
            row,
            10,
            investment.profit_loss
        )

        investment_sheet.cell(
            row,
            11,
            round(
                investment.return_percentage,
                2
            )
        )

    # --------------------------------------------------------
    # Auto Width
    # --------------------------------------------------------

    for sheet in workbook.worksheets:

        for column in sheet.columns:

            max_length = 0

            for cell in column:

                if cell.value is not None:

                    max_length = max(

                        max_length,

                        len(
                            str(cell.value)
                        )
                    )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = (
                max_length + 2
            )

    # --------------------------------------------------------
    # Generate Excel File
    # --------------------------------------------------------

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            "FinSight_Financial_Report.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )
    )


# ============================================================
# PDF FINANCIAL REPORT
# ============================================================


# ============================================================
# PDF REPORT PAGE
# ============================================================

@app.route("/pdf_report")
@login_required
def pdf_report():

    return render_template(
        "pdf_report.html"
    )


# ============================================================
# DOWNLOAD PDF
# ============================================================

@app.route("/download_pdf")
@login_required
def download_pdf():

    # --------------------------------------------------------
    # Get current user's data
    # --------------------------------------------------------

    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()

    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()

    goals = FinancialGoal.query.filter_by(
        user_id=current_user.id
    ).all()

    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()

    # --------------------------------------------------------
    # Calculate summary
    # --------------------------------------------------------

    total_expenses = sum(
        expense.amount
        for expense in expenses
    )

    total_budget = sum(
        budget.amount
        for budget in budgets
    )

    remaining_balance = (
        total_budget -
        total_expenses
    )

    total_invested = sum(
        investment.invested_amount
        for investment in investments
    )

    total_investment_value = sum(
        investment.current_value
        for investment in investments
    )

    total_profit_loss = (
        total_investment_value -
        total_invested
    )

    # --------------------------------------------------------
    # Create PDF
    # --------------------------------------------------------

    output = BytesIO()

    document = SimpleDocTemplate(

        output,

        pagesize=A4,

        rightMargin=40,

        leftMargin=40,

        topMargin=40,

        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    content = []

    # --------------------------------------------------------
    # Title
    # --------------------------------------------------------

    content.append(

        Paragraph(

            "FinSight - Financial Report",

            styles["Title"]
        )
    )

    content.append(
        Spacer(1, 15)
    )

    content.append(

        Paragraph(

            f"User: {current_user.username}",

            styles["Normal"]
        )
    )

    content.append(
        Spacer(1, 20)
    )

    # ========================================================
    # EXPENSES
    # ========================================================

    content.append(

        Paragraph(
            "1. Expenses",
            styles["Heading2"]
        )
    )

    expense_data = [[

        "Title",
        "Category",
        "Amount",
        "Date"
    ]]

    for expense in expenses:

        expense_data.append([

            str(
                expense.title
            ),

            str(
                expense.category
            ),

            f"Rs. {expense.amount:.2f}",

            str(
                expense.date
            )
        ])

    if len(expense_data) == 1:

        expense_data.append([

            "No expenses",
            "-",
            "-",
            "-"
        ])

    expense_data.append([

        "",
        "",

        f"Total: Rs. {total_expenses:.2f}",

        ""
    ])

    expense_table = Table(

        expense_data,

        colWidths=[
            130,
            100,
            100,
            100
        ],

        repeatRows=1
    )

    expense_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (2, 1),
                (2, -1),
                "RIGHT"
            )
        ])
    )

    content.append(
        expense_table
    )

    content.append(
        Spacer(1, 20)
    )

    # ========================================================
    # BUDGETS
    # ========================================================

    content.append(

        Paragraph(
            "2. Budgets",
            styles["Heading2"]
        )
    )

    budget_data = [[

        "Month",
        "Budget Amount"
    ]]

    for budget in budgets:

        budget_data.append([

            str(
                budget.month
            ),

            f"Rs. {float(budget.amount):.2f}"
        ])

    if len(budget_data) == 1:

        budget_data.append([

            "No budgets",
            "-"
        ])

    budget_data.append([

        "Total",

        f"Rs. {total_budget:.2f}"
    ])

    budget_table = Table(

        budget_data,

        colWidths=[
            200,
            200
        ],

        repeatRows=1
    )

    budget_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            )
        ])
    )

    content.append(
        budget_table
    )

    content.append(
        Spacer(1, 20)
    )

    # ========================================================
    # FINANCIAL GOALS
    # ========================================================

    content.append(

        Paragraph(
            "3. Financial Goals",
            styles["Heading2"]
        )
    )

    goal_data = [[

        "Goal",
        "Target",
        "Current",
        "Target Date"
    ]]

    for goal in goals:

        goal_data.append([

            str(
                goal.goal_name
            ),

            f"Rs. {goal.target_amount:.2f}",

            f"Rs. {goal.current_amount:.2f}",

            str(
                goal.target_date
            )
        ])

    if len(goal_data) == 1:

        goal_data.append([

            "No goals",
            "-",
            "-",
            "-"
        ])

    goal_table = Table(

        goal_data,

        colWidths=[
            130,
            100,
            100,
            100
        ],

        repeatRows=1
    )

    goal_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            )
        ])
    )

    content.append(
        goal_table
    )

    content.append(
        Spacer(1, 20)
    )

    # ========================================================
    # INVESTMENTS
    # ========================================================

    content.append(

        Paragraph(
            "4. Investments",
            styles["Heading2"]
        )
    )

    investment_data = [[

        "Asset",
        "Type",
        "Qty",
        "Purchase Price",
        "Current Price"
    ]]

    for investment in investments:

        investment_data.append([

            str(
                investment.asset_name
            ),

            str(
                investment.asset_type
            ),

            str(
                investment.quantity
            ),

            f"Rs. {investment.purchase_price:.2f}",

            f"Rs. {investment.current_price:.2f}"
        ])

    if len(investment_data) == 1:

        investment_data.append([

            "No investments",
            "-",
            "-",
            "-",
            "-"
        ])

    investment_table = Table(

        investment_data,

        colWidths=[
            90,
            80,
            55,
            105,
            105
        ],

        repeatRows=1
    )

    investment_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (2, 1),
                (2, -1),
                "RIGHT"
            )
        ])
    )

    content.append(
        investment_table
    )

    content.append(
        Spacer(1, 20)
    )

    # ========================================================
    # FINANCIAL SUMMARY
    # ========================================================

    content.append(

        Paragraph(
            "Financial Summary",
            styles["Heading2"]
        )
    )

    summary_data = [

        [
            "Total Budget",
            f"Rs. {total_budget:.2f}"
        ],

        [
            "Total Expenses",
            f"Rs. {total_expenses:.2f}"
        ],

        [
            "Remaining Balance",
            f"Rs. {remaining_balance:.2f}"
        ],

        [
            "Total Invested",
            f"Rs. {total_invested:.2f}"
        ],

        [
            "Current Investment Value",
            f"Rs. {total_investment_value:.2f}"
        ],

        [
            "Investment Profit/Loss",
            f"Rs. {total_profit_loss:.2f}"
        ]
    ]

    summary_table = Table(

        summary_data,

        colWidths=[
            220,
            220
        ]
    )

    summary_table.setStyle(

        TableStyle([

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (0, -1),
                "Helvetica-Bold"
            )
        ])
    )

    content.append(
        summary_table
    )

    # --------------------------------------------------------
    # Build PDF
    # --------------------------------------------------------

    document.build(
        content
    )

    output.seek(0)

    # ========================================================
    # DOWNLOAD PDF
    # ========================================================

    return send_file(

        output,

        as_attachment=True,

        download_name="FinSight_Financial_Report.pdf",

        mimetype="application/pdf"
    )
# ============================================================
# LOGOUT
# ============================================================

@app.route("/logout")
@login_required
def logout():

    logout_user()

    return redirect(
        url_for("login")
    )


# ============================================================
# RUN APP
# ============================================================

if __name__ == "__main__":

    with app.app_context():

        db.create_all()

    app.run(
        debug=True
    )